import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet('2010')
ws.append(['기간','순위','프로그램','시청률'])
for weekIndex in range(0,5):
    res = requests.get("https://workey.codeit.kr/ratings/index?year=2010&month=1&weekIndex={}".format(weekIndex))
    rating_page = res.text
    # HTML 정리
    soup = BeautifulSoup(rating_page, 'html.parser')
    for tag in soup.select('tr')[1:]:
        td_tags = tag.select('td')
        if td_tags[1].get_text() == 'SBS':

            proid = "2020년 1월 {}주차".format(weekIndex+1)
            row = [
                proid,
                td_tags[0].get_text(),
                td_tags[1].get_text(),
                td_tags[2].get_text(),
                td_tags[3].get_text()
            ]
            ws.append(row)
wb.save('test.xlsx')

# for year in range(2010, 2019):
#     ws = wb.create_sheet("{}년".format(year))
#     ws.append(['순위','채널','프로그램','시청률'])
#     for month in range(1,13):
#         for weekIndex in range(0,5):
#             res = requests.get("https://workey.codeit.kr/ratings/index?year={}&month={}&weekIndex={}".format(year, month, weekIndex))
#             rating_page = res.text
#             # HTML 정리
#             soup = BeautifulSoup(rating_page, 'html.parser')
#             for tag in soup.select('tr')[1:]:
#                 td_tags = tag.select('td')
#                 row = [
#                     td_tags[0].get_text(),
#                     td_tags[1].get_text(),
#                     td_tags[2].get_text(),
#                     td_tags[3].get_text(),
#                 ]
#                 ws.append(row)
#
# wb.save('시청률s.xlsx')
